import math
from collections import defaultdict
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.styles import Font, Color
from sklearn.utils import shuffle
from icr_lib import AveragingClassifier, preprocess_data, read_and_prepreprocess_data, fit_cv, get_params_dict, \
    get_pipeline, GpuQueue, update_params_dict


def save_results_to_excel(df, filename):
    filename = filename if filename.endswith('.xlsx') else filename + '.xlsx'
    df.to_excel(filename, index=False)

    # Open the file for editing
    workbook = load_workbook(filename)

    # Select the sheet
    sheet = workbook.active

    # Highlight the Log Loss columns
    red_font = Font(color=Color(rgb="FF0000"))
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "Log Loss" in cell.value:
                cell.font = red_font

    workbook.save(filename)


def aggregate_results(rounds_metrics):
    df = pd.DataFrame(rounds_metrics)
    df.to_csv('Results/rounds_metrics.csv', index=False)
    save_results_to_excel(df, 'rounds_metrics.xlsx')

    df_final = df.groupby(by=['model_name', 'metric_name', 'measure_type']).mean().reset_index().drop(columns='round')
    df_final.to_csv('Results/final_metrics.csv', index=False)
    save_results_to_excel(df_final, 'final_metrics.xlsx')
    i = iter(iter_subplots_axes(1, 1, 16, 9))
    next(i)
    plt.title('cross validation measurements')
    sns.barplot(data=df[df.measure_type == 'cv'], hue='model_name', x='metric_name', y='metric_value')
    plt.grid()
    plt.savefig('stats.png')


def iter_subplots_axes(ncol, n_subplots, tile_size_col=5, tile_size_row=5, title=None, title_fontsize=14):
    """ Creates subplots figure, and iterates over axes in left-right/top-bottom order """
    nrow = math.ceil(n_subplots / ncol)
    fig, axes = plt.subplots(nrow, ncol)
    if title is not None:
        plt.suptitle(title, fontsize=title_fontsize)
    fig.set_size_inches(ncol * tile_size_col, nrow * tile_size_row)
    for i in range(n_subplots):
        if nrow > 1 and ncol > 1:
            ax = axes[i // ncol, i % ncol]
        else:
            if n_subplots > 1:
                ax = axes[i]
            else:
                ax = axes
        plt.sca(ax)
        yield ax


def try_models(X, y, n_rounds, models, params):
    classifier = get_pipeline(models, params, X, y, GpuQueue())

    # create an empty list to store the metrics
    metrics_rows = []
    final_model_list = []

    cv_score_means = defaultdict(list)
    # for round in range(n_rounds):
    for round in range(n_rounds):
        X, y = shuffle(X, y, random_state=round)
        classifier.fit(X, y)
        final_model_list.append(classifier)
        cross_val_scores, test_scores_current_round = fit_cv(classifier, round, X, y)
        for cv_scores in cross_val_scores:
            for metric, values in cv_scores.items():
                cv_score_means[metric].append(np.mean(values))
        for metric, values in cv_score_means.items():
            metrics_rows.append(dict(model_name='optuna_pipeline', round=round,
                                     metric_name=metric, metric_value=np.mean(values),
                                     measure_type='cv'))
    aggregate_results(metrics_rows)
    return final_model_list


def main():
    X, y, _ = read_and_prepreprocess_data()
    models = ['catboost', 'xgboost', 'xgboost2', 'xgboost3', 'lgbm', 'tabpfn24', 'tabpfn64']
    params = {'dimensionality_reduction_model1_index': 4, 'dimensionality_reduction_model2_index': 3,
              'catboost_iterations': 119, 'catboost_learning_rate': 0.2373478627191737, 'catboost_depth': 4,
              'catboost_l2_leaf_reg': 1.3063393283179154, 'xgboost_n_estimators': 165, 'xgboost_max_depth': 3,
              'xgboost_subsample': 0.8119794123653089, 'xgboost_colsample_bytree': 0.9438318190982824,
              'lgbm_max_depth': 5, 'lgbm_learning_rate': 0.2785788274064489, 'lgbm_num_leaves': 120,
              'lgbm_colsample_bytree': 0.9729778631903917, 'xgboost2_n_estimators': 191, 'xgboost2_max_depth': 7,
              'xgboost2_subsample': 0.7330530954570363, 'xgboost2_colsample_bytree': 0.6394804662463566,
              'xgboost3_n_estimators': 119, 'xgboost3_max_depth': 8, 'xgboost3_subsample': 0.9671996924315035,
              'xgboost3_colsample_bytree': 0.6744071411442031, 'AB': True, 'AF': False, 'AH': True, 'AM': False,
              'AR': False, 'AX': False, 'AY': False, 'AZ': False, 'BC': True, 'BD ': False, 'BN': False, 'BP': False,
              'BQ': True, 'BR': False, 'BZ': False, 'CB': False, 'CC': True, 'CD ': False, 'CF': True, 'CH': True,
              'CL': False, 'CR': False, 'CS': False, 'CU': False, 'CW ': False, 'DA': False, 'DE': False, 'DF': False,
              'DH': False, 'DI': False, 'DL': True, 'DN': True, 'DU': True, 'DV': True, 'DY': True, 'EB': True,
              'EE': True, 'EG': True, 'EH': True, 'EJ': False, 'EL': True, 'EP': True, 'EU': True, 'FC': False,
              'FD ': True, 'FE': False, 'FI': False, 'FL': False, 'FR': False, 'FS': True, 'GB': True, 'GE': True,
              'GF': False, 'GH': False, 'GI': True, 'GL': True, 'Epsilon': False, 'BQ_is_nan': False,
              'oversampling_catboost': False, 'oversampling_xgboost': True, 'oversampling_xgboost2': False,
              'oversampling_xgboost3': True, 'oversampling_lgbm': False}
    params = update_params_dict(X, models, params)
    print('updated params dict: ', params)
    final_model_list = try_models(X, y, n_rounds=1, models=models, params=params)
    X_test, _, final_df = read_and_prepreprocess_data(for_training=False)
    assert len(X_test) > 0

    model_probabilities_list = []
    for model in final_model_list:
        model_probabilities_list.append(model.predict_proba(X_test))
    final_probabilities = np.mean(model_probabilities_list, axis=0)
    final_probabilities_df = pd.DataFrame(final_probabilities, columns=['class_0', 'class_1'])
    final_probabilities_df['Id'] = final_df['Id']
    final_df = final_probabilities_df[['Id', 'class_0', 'class_1']]
    submission_csv = final_df.to_csv('submission.csv', index=False)


if __name__ == '__main__':
    main()

# /kaggle/input/icr-identify-age-related-conditions/